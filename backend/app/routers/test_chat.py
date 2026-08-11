import asyncio
import json
import logging
import re
from typing import Optional, List
from io import BytesIO
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.sql import func
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from app.database import get_db, SessionLocal
from app.models.user import User
from app.models.agent import Agent
from app.models.workflow_agent import WorkflowAgent
from app.models.functional_model import FunctionalModel
from app.models.test_case_folder import TestCaseFolder
from app.models.test_case import TestCase
from app.models.conversation import Conversation, TestConversationState, ChatMessage, TestResult
from app.models.file import BaseFile
from app.models.test_case_image import TestCaseImage
from app.models.extracted_image import ExtractedImage
from app.models.api_config import ApiConfig
from app.auth import get_current_user
from app.services.conversation_saver import save_conversation_messages
from app.services.file_storage import file_storage
from app.services.chat_helpers import get_chat_config, prepare_messages, build_system_prompt_with_context
from app.routers.test_executor import run_tests_for_multiple_agents
from app.websocket_manager import manager
from app.services.llm_client import (
    get_intent_model_config,
    call_llm,
    stream_llm
)
from app.prompts import (
    TEST_INTENT_SYSTEM_PROMPT,
    TEST_INTENT_USER_PROMPT,
    CONTINUE_TEST_SYSTEM_PROMPT,
    CONTINUE_TEST_USER_PROMPT,
    TEST_CASE_GENERATION_SYSTEM_PROMPT,
    TEST_CASE_GENERATION_USER_PROMPT
)
import httpx
import uuid
import time

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/test-chat", tags=["测试智能体模式聊天"])


class TestChatRequest(BaseModel):
    messages: List[dict]
    conversation_id: Optional[int] = None
    conversation_context: Optional[str] = ""
    current_status: Optional[str] = "pending"
    current_agent_names: Optional[List[str]] = []
    current_test_case: Optional[str] = ""
    current_request_interval: Optional[int] = 0
    enable_thinking: Optional[bool] = False
    enable_search: Optional[bool] = False
    config_id: Optional[int] = None


class TestIntentResponse(BaseModel):
    test_case: str
    agent_names: List[str]
    status: str
    message: str
    request_interval: int = 0
    conversation_id: Optional[int] = None
    message_id: Optional[int] = None


class GenerateTestCaseRequest(BaseModel):
    agent_names: List[str]
    requirement: Optional[str] = ""
    count: Optional[int] = 5
    folder_id: Optional[str] = None


class GeneratedTestCase(BaseModel):
    question: str
    sample_answer: str


class GenerateTestCaseResponse(BaseModel):
    test_cases: List[GeneratedTestCase]
    agent_names: List[str]


def get_test_results_info(db: Session, conversation_id: int) -> str:
    results = db.query(TestResult).filter(
        TestResult.conversation_id == conversation_id
    ).all()
    
    if not results:
        return "（暂无测试结果）"
    
    result_lines = []
    for result in results:
        agent = db.query(Agent).filter(Agent.id == result.agent_id).first()
        workflow_agent = db.query(WorkflowAgent).filter(WorkflowAgent.id == result.agent_id).first()
        if agent:
            agent_name = agent.name
        else: 
            agent_name = workflow_agent.name if workflow_agent else "未知智能体"
        
        test_case = db.query(TestCase).filter(TestCase.id == result.test_case_id).first()
        sample_answer = test_case.sample_answer if test_case else ""
        
        result_lines.append(f"### 智能体: {agent_name}")
        result_lines.append(f"**问题**: {result.question or '无'}")
        result_lines.append(f"**智能体回答**: {result.response or '无回答'}")
        if sample_answer:
            result_lines.append(f"**参考答案**: {sample_answer}")
        if result.error_message:
            result_lines.append(f"**错误信息**: {result.error_message}")
        result_lines.append(f"**请求耗时**: {result.request_time:.2f}秒" if result.request_time else "**请求耗时**: 未知")
        result_lines.append("")
    
    return "\n".join(result_lines)


def get_agents_info(db: Session, user_id: int) -> str:
    agents = db.query(Agent).filter(
        Agent.user_id == user_id,
        Agent.is_active == True
    ).all()
    
    workflow_agents = db.query(WorkflowAgent).filter(
        WorkflowAgent.user_id == user_id,
        WorkflowAgent.is_active == True
    ).all()
    
    if not agents and not workflow_agents:
        return "（暂无可用智能体）"
    
    lines = []
    
    if agents:
        lines.append("【普通智能体】")
        for agent in agents:
            lines.append(f"- {agent.name}")
    
    if workflow_agents:
        lines.append("【工作流智能体】")
        for agent in workflow_agents:
            lines.append(f"- {agent.name}")
    
    return "\n".join(lines)


def get_test_folders_info(db: Session, user_id: int) -> str:
    folders = db.query(TestCaseFolder).filter(
        TestCaseFolder.user_id == user_id
    ).all()
    
    if not folders:
        return "（暂无可用测试用例文件夹）"
    
    return "\n".join([f"- {folder.name}" for folder in folders])


def validate_agent_names(db: Session, user_id: int, agent_names: List[str]) -> List[dict]:
    valid_agents = []
    
    agents = db.query(Agent).filter(
        Agent.user_id == user_id,
        Agent.is_active == True
    ).all()
    valid_agent_names = {agent.name: agent.id for agent in agents}
    
    workflow_agents = db.query(WorkflowAgent).filter(
        WorkflowAgent.user_id == user_id,
        WorkflowAgent.is_active == True
    ).all()
    valid_workflow_agent_names = {agent.name: agent.id for agent in workflow_agents}
    
    for name in agent_names:
        if name in valid_agent_names:
            valid_agents.append({
                "name": name,
                "id": valid_agent_names[name],
                "type": "agent"
            })
        elif name in valid_workflow_agent_names:
            valid_agents.append({
                "name": name,
                "id": valid_workflow_agent_names[name],
                "type": "workflow_agent"
            })
    
    return valid_agents


def validate_test_case(db: Session, user_id: int, test_case: str) -> Optional[str]:
    if not test_case:
        return None
    
    folder = db.query(TestCaseFolder).filter(
        TestCaseFolder.user_id == user_id,
        TestCaseFolder.name == test_case
    ).first()
    
    return folder.name if folder else None


def save_test_state(
    db: Session,
    conversation_id: int,
    status: str,
    agent_names,
    test_case: str,
    request_interval: int = 0
):
    if agent_names and isinstance(agent_names[0], dict):
        names_list = [a["name"] for a in agent_names]
    else:
        names_list = agent_names
    
    test_state = db.query(TestConversationState).filter(
        TestConversationState.conversation_id == conversation_id
    ).first()
    
    if test_state:
        test_state.status = status
        test_state.agent_names = ','.join(names_list) if names_list else ""
        test_state.test_case = test_case
        test_state.request_interval = request_interval
    else:
        test_state = TestConversationState(
            conversation_id=conversation_id,
            status=status,
            agent_names=','.join(names_list) if names_list else "",
            test_case=test_case,
            request_interval=request_interval
        )
        db.add(test_state)
    
    db.commit()
    return test_state


def save_test_messages(
    db: Session,
    request,
    current_user: User,
    user_message: str,
    assistant_message: str,
    agent_names: List[str],
    test_case: str,
    status: str,
    request_interval: int = 0
) -> tuple:
    conversation_id, msg_id = save_conversation_messages(
        db=db,
        user_id=current_user.id,
        conversation_id=request.conversation_id,
        user_message=user_message,
        assistant_message=assistant_message,
        api_id=request.config_id,
        conversation_mode="test"
    )
    
    save_test_state(
        db=db,
        conversation_id=conversation_id,
        status=status,
        agent_names=agent_names,
        test_case=test_case,
        request_interval=request_interval
    )
    
    return conversation_id, msg_id


@router.post("/recognize-intent", response_model=None)
async def recognize_test_intent(
    request: TestChatRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    messages = request.messages
    last_message = messages[-1]["content"] if messages else ""
    
    if request.conversation_id:
        conversation = db.query(Conversation).filter(
            Conversation.id == request.conversation_id,
            Conversation.user_id == current_user.id
        ).first()
        
        if conversation and conversation.test_state:
            current_status = conversation.test_state.status
            current_agent_names = [name.strip() for name in conversation.test_state.agent_names.split(',') if name.strip()]
            current_test_case = conversation.test_state.test_case
            current_request_interval = conversation.test_state.request_interval
        else:
            current_status = "pending"
            current_agent_names = []
            current_test_case = ""
            current_request_interval = 0
    else:
        current_status = request.current_status or "pending"
        current_agent_names = request.current_agent_names or []
        current_test_case = request.current_test_case or ""
        current_request_interval = request.current_request_interval or 0
    
    logger.info(f"=== Test Chat Request ===")
    logger.info(f"conversation_id: {request.conversation_id}")
    logger.info(f"current_status: {current_status}")
    logger.info(f"last_message: {last_message}")
    logger.info(f"current_agent_names: {current_agent_names}")
    logger.info(f"current_test_case: {current_test_case}")
    logger.info(f"current_request_interval: {current_request_interval}")
    
    agent_names_from_at = re.findall(r'@([^\s@]+)', last_message)
    
    agents_info = get_agents_info(db, current_user.id)
    test_folders_info = get_test_folders_info(db, current_user.id)
    
    intent_model_config = get_intent_model_config(db)
    
    if not intent_model_config:
        logger.warning("No intent recognition model configured")
        valid_agents = validate_agent_names(db, current_user.id, agent_names_from_at)
        
        conv_id, msg_id = save_test_messages(
            db=db,
            request=request,
            current_user=current_user,
            user_message=last_message,
            assistant_message="请配置意图识别模型，或手动指定智能体和测试用例",
            agent_names=[a["name"] for a in valid_agents],
            test_case="",
            status="pending",
            request_interval=current_request_interval
        )
        
        return TestIntentResponse(
            test_case="",
            agent_names=[a["name"] for a in valid_agents],
            status="pending",
            message="请配置意图识别模型，或手动指定智能体和测试用例",
            request_interval=current_request_interval,
            conversation_id=conv_id,
            message_id=msg_id
        )
    
    if current_status == "completed":
        logger.info(f"=== Handling completed status ===")
        
        continue_test_messages = [
            {"role": "system", "content": CONTINUE_TEST_SYSTEM_PROMPT},
            {"role": "user", "content": CONTINUE_TEST_USER_PROMPT.format(
                message=last_message,
                agents_info=agents_info,
                test_folders_info=test_folders_info
            )}
        ]
        
        try:
            continue_test_result = await call_llm(
                messages=continue_test_messages,
                model=intent_model_config["model"],
                api_key=intent_model_config["api_key"],
                url=intent_model_config["url"],
                call_type=intent_model_config.get("call_type", "OpenAI Chat"),
                structured_output=True,
                timeout=30.0,
                default_response={"continue_test": False}
            )
            continue_test = continue_test_result.get("continue_test", False)
            logger.info(f"Continue test check result: {continue_test}")
        except Exception as e:
            logger.error(f"Continue test check failed: {str(e)}")
            continue_test = False
        
        if continue_test:
            logger.info("User wants to continue testing, treating as pending status")
            current_status = "pending"
            current_agent_names = []
            current_test_case = ""
        else:
            logger.info("User does not want to continue testing, will stream response")
            
            if not request.config_id:
                conv_id, msg_id = save_test_messages(
                    db=db,
                    request=request,
                    current_user=current_user,
                    user_message=last_message,
                    assistant_message="测试已完成。如需继续测试其他智能体，请告诉我。",
                    agent_names=[],
                    test_case="",
                    status="completed",
                    request_interval=current_request_interval
                )
                return TestIntentResponse(
                    test_case="",
                    agent_names=[],
                    status="completed",
                    message="测试已完成。如需继续测试其他智能体，请告诉我。",
                    request_interval=current_request_interval,
                    conversation_id=conv_id,
                    message_id=msg_id
                )
            
            try:
                config = get_chat_config(db, request.config_id, current_user.id)
            except HTTPException:
                conv_id, msg_id = save_test_messages(
                    db=db,
                    request=request,
                    current_user=current_user,
                    user_message=last_message,
                    assistant_message="测试已完成。模型配置不存在，如需继续测试其他智能体，请告诉我。",
                    agent_names=[],
                    test_case="",
                    status="completed",
                    request_interval=current_request_interval
                )
                return TestIntentResponse(
                    test_case="",
                    agent_names=[],
                    status="completed",
                    message="测试已完成。模型配置不存在，如需继续测试其他智能体，请告诉我。",
                    request_interval=current_request_interval,
                    conversation_id=conv_id,
                    message_id=msg_id
                )
            
            test_results_info = ""
            if request.conversation_id:
                test_results_info = get_test_results_info(db, request.conversation_id)
            
            system_prompt = build_system_prompt_with_context(
                test_results_info,
                "以下是刚才的测试结果，请根据用户的问题进行分析和回答："
            )
            
            chat_messages = [{"role": msg["role"], "content": msg["content"]} for msg in messages]
            chat_messages.insert(0, {"role": "system", "content": system_prompt})
            
            async def test_chat_response_stream():
                full_content = ""
                async for chunk in stream_llm(
                    messages=chat_messages,
                    model=config.model_code,
                    api_key=config.api_key,
                    url=config.url,
                    call_type=config.call_type,
                    enable_thinking=request.enable_thinking or False,
                    timeout=60.0,
                    enable_search=request.enable_search or False
                ):
                    if chunk.startswith("data: "):
                        data = chunk[6:]
                        if data == "[DONE]":
                            break
                        try:
                            parsed = json.loads(data)
                            if "content" in parsed:
                                full_content += parsed["content"]
                        except (json.JSONDecodeError, TypeError):
                            pass
                    yield chunk
                
                conversation_id = request.conversation_id
                
                if not conversation_id:
                    new_conversation = Conversation(
                        user_id=current_user.id,
                        title=last_message[:50] if len(last_message) > 50 else last_message,
                        conversation_mode="test",
                        api_id=request.config_id
                    )
                    db.add(new_conversation)
                    db.flush()
                    conversation_id = new_conversation.id
                
                user_message = ChatMessage(
                    conversation_id=conversation_id,
                    role="user",
                    content=last_message
                )
                db.add(user_message)
                db.flush()
                
                assistant_message = ChatMessage(
                    conversation_id=conversation_id,
                    role="assistant",
                    content=full_content
                )
                db.add(assistant_message)
                db.flush()
                
                db.commit()
                
                yield f"data: {json.dumps({'type': 'saved', 'conversation_id': conversation_id, 'message_id': assistant_message.id})}\n\n"
                yield "data: [DONE]\n\n"
            
            return StreamingResponse(test_chat_response_stream(), media_type="text/event-stream")

    if current_status in ["pending", "lacking_test_case", "lacking_agent_name", "success_generating", "continue_test", "start_testing"]:
        logger.info(f"=== Handling {current_status} status with unified intent recognition ===")
        
        # 统一使用同一个意图识别提示词
        intent_messages = [
            {"role": "system", "content": TEST_INTENT_SYSTEM_PROMPT},
            {"role": "user", "content": TEST_INTENT_USER_PROMPT.format(
                message=last_message + "\n历史对话：\n" + "\n".join([msg["content"] for msg in messages[:-1]]),
                agents_info=agents_info,
                test_folders_info=test_folders_info
            )}
        ]
        
        try:
            intent_result = await call_llm(
                messages=intent_messages,
                model=intent_model_config["model"],
                api_key=intent_model_config["api_key"],
                url=intent_model_config["url"],
                call_type=intent_model_config.get("call_type", "OpenAI Chat"),
                structured_output=True,
                timeout=30.0,
                default_response={"motion": "unassigned", "agent_names": [], "test_case": "", "requirement": ""}
            )
            logger.info(f"Intent recognition result: {intent_result}")
        except Exception as e:
            logger.error(f"Intent recognition failed: {str(e)}")
            intent_result = {"motion": "unassigned", "agent_names": [], "test_case": "", "requirement": ""}
        
        # 处理生成测试用例的意图
        if intent_result.get("motion") == "agent_generating_test_case":
            logger.info(f"User requested to generate test case via intent recognition")
            
            extracted_agent_names_raw = intent_result.get("agent_names", [])
            extracted_requirement = intent_result.get("requirement", last_message)
            
            if not extracted_agent_names_raw:
                extracted_agents = []
                for name in current_agent_names:
                    agents = validate_agent_names(db, current_user.id, [name])
                    extracted_agents.extend(agents)
            else:
                extracted_agents = validate_agent_names(db, current_user.id, extracted_agent_names_raw)
            
            if not extracted_agents:
                conv_id, msg_id = save_test_messages(
                    db=db,
                    request=request,
                    current_user=current_user,
                    user_message=last_message,
                    assistant_message=f"请先指定要测试的智能体名称（使用 @智能体名称 格式）。\n\n可用的智能体：\n{agents_info}",
                    agent_names=[],
                    test_case="",
                    status="lacking_agent_name",
                    request_interval=current_request_interval
                )
                
                return TestIntentResponse(
                    test_case="",
                    agent_names=[],
                    status="lacking_agent_name",
                    message=f"请先指定要测试的智能体名称（使用 @智能体名称 格式）。\n\n可用的智能体：\n{agents_info}",
                    request_interval=current_request_interval,
                    conversation_id=conv_id,
                    message_id=msg_id
                )
            
            extracted_agent_names = [a["name"] for a in extracted_agents]
            
            background_tasks.add_task(
                generate_test_case_background,
                agent_names=extracted_agent_names,
                user_id=current_user.id,
                user_requirement=extracted_requirement,
                count=5,
                conversation_id=request.conversation_id
            )
            
            conv_id, msg_id = save_test_messages(
                db=db,
                request=request,
                current_user=current_user,
                user_message=last_message,
                assistant_message=f"将为智能体 **{', '.join(extracted_agent_names)}** 自动生成测试用例。\n\n请稍候...",
                agent_names=extracted_agent_names,
                test_case="",
                status="agent_generate_test_case",
                request_interval=current_request_interval
            )
            
            return TestIntentResponse(
                test_case="",
                agent_names=extracted_agent_names,
                status="agent_generate_test_case",
                message=f"将为智能体 **{', '.join(extracted_agent_names)}** 自动生成测试用例。\n\n请稍候...",
                request_interval=current_request_interval,
                conversation_id=conv_id,
                message_id=msg_id
            )
        
        # 处理指定测试用例的情况
        extracted_test_case = intent_result.get("test_case", "")
        valid_test_case = validate_test_case(db, current_user.id, extracted_test_case)
        
        if not valid_test_case:
            # 尝试直接验证用户输入
            valid_test_case = validate_test_case(db, current_user.id, last_message.strip())
        
        # 处理指定智能体的情况
        extracted_agent_names_raw = intent_result.get("agent_names", [])
        
        if not extracted_agent_names_raw:
            agent_names_from_at = re.findall(r'@([^\s@]+)', last_message)
            extracted_agents = validate_agent_names(db, current_user.id, agent_names_from_at)
        else:
            extracted_agents = validate_agent_names(db, current_user.id, extracted_agent_names_raw)
        
        extracted_agent_names = [a["name"] for a in extracted_agents]
        
        final_agent_names = current_agent_names.copy()
        for agent_name in extracted_agent_names:
            if agent_name not in final_agent_names:
                final_agent_names.append(agent_name)
        
        if valid_test_case:
            if final_agent_names:
                status = "start_testing"
                message = f"测试信息已确认。\n\n**智能体**：{', '.join(final_agent_names)}\n**测试用例**：{valid_test_case}\n"
            else:
                status = "lacking_agent_name"
                message = f"测试用例已确认：{valid_test_case}\n\n请指定要测试的智能体名称（使用 @智能体名称 格式）。\n\n可用的智能体：\n{agents_info}"
            
            logger.info(f"Test case provided - status: {status}")
            
            conv_id, msg_id = save_test_messages(
                db=db,
                request=request,
                current_user=current_user,
                user_message=last_message,
                assistant_message=message,
                agent_names=final_agent_names,
                test_case=valid_test_case,
                status=status,
                request_interval=current_request_interval
            )
            
            return TestIntentResponse(
                test_case=valid_test_case,
                agent_names=final_agent_names,
                status=status,
                message=message,
                request_interval=current_request_interval,
                conversation_id=conv_id,
                message_id=msg_id
            )
        elif final_agent_names:
            status = "lacking_test_case"
            message = f"智能体已确认：{', '.join(final_agent_names)}\n\n请提供测试用例文件夹名称，或回复「生成」让智能体自动生成测试用例。\n\n可用的测试用例文件夹：\n{test_folders_info}"
            
            conv_id, msg_id = save_test_messages(
                db=db,
                request=request,
                current_user=current_user,
                user_message=last_message,
                assistant_message=message,
                agent_names=final_agent_names,
                test_case="",
                status=status,
                request_interval=current_request_interval
            )
            
            return TestIntentResponse(
                test_case="",
                agent_names=final_agent_names,
                status=status,
                message=message,
                request_interval=current_request_interval,
                conversation_id=conv_id,
                message_id=msg_id
            )
        else:
            # 既没有智能体也没有测试用例
            status = "pending"
            message = f"请指定要测试的智能体（使用 @智能体名称 格式）和测试用例。\n\n可用的智能体：\n{agents_info}\n\n可用的测试用例文件夹：\n{test_folders_info}"
            
            conv_id, msg_id = save_test_messages(
                db=db,
                request=request,
                current_user=current_user,
                user_message=last_message,
                assistant_message=message,
                agent_names=[],
                test_case="",
                status=status,
                request_interval=current_request_interval
            )
            
            return TestIntentResponse(
                test_case="",
                agent_names=[],
                status=status,
                message=message,
                request_interval=current_request_interval,
                conversation_id=conv_id,
                message_id=msg_id
            )
    
    
    
    # 处理其他状态
    logger.info(f"=== Handling default status: {current_status} ===")
    
    # 默认为pending状态
    status = "pending"
    message = f"请指定要测试的智能体（使用 @智能体名称 格式）和测试用例。\n\n可用的智能体：\n{agents_info}\n\n可用的测试用例文件夹：\n{test_folders_info}"
    valid_agent_names = []
    valid_test_case = ""
    
    logger.info(f"=== Test Chat Response ===")
    logger.info(f"status: {status}")
    logger.info(f"message: {message}")
    logger.info(f"agent_names: {valid_agent_names}")
    logger.info(f"test_case: {valid_test_case or ''}")
    
    conv_id, msg_id = save_test_messages(
        db=db,
        request=request,
        current_user=current_user,
        user_message=last_message,
        assistant_message=message,
        agent_names=valid_agent_names,
        test_case=valid_test_case or "",
        status=status,
        request_interval=current_request_interval
    )
    
    return TestIntentResponse(
        test_case=valid_test_case or "",
        agent_names=valid_agent_names,
        status=status,
        message=message,
        request_interval=current_request_interval,
        conversation_id=conv_id,
        message_id=msg_id
    )


class UpdateIntervalRequest(BaseModel):
    conversation_id: int
    request_interval: int


@router.post("/update-interval")
async def update_request_interval(
    request: UpdateIntervalRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    conversation = db.query(Conversation).filter(
        Conversation.id == request.conversation_id,
        Conversation.user_id == current_user.id
    ).first()
    
    if not conversation:
        raise HTTPException(status_code=404, detail="对话不存在")
    
    test_state = db.query(TestConversationState).filter(
        TestConversationState.conversation_id == request.conversation_id
    ).first()
    
    if test_state:
        test_state.request_interval = request.request_interval
    else:
        test_state = TestConversationState(
            conversation_id=request.conversation_id,
            status="pending",
            agent_names="",
            test_case="",
            request_interval=request.request_interval
        )
        db.add(test_state)
    
    db.commit()
    
    return {"success": True, "request_interval": request.request_interval}


@router.get("/agents")
async def get_agents_list(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    agents = db.query(Agent).filter(
        Agent.user_id == current_user.id,
        Agent.is_active == True
    ).all()
    
    return [{"id": agent.id, "name": agent.name} for agent in agents]


@router.get("/test-folders")
async def get_test_folders_list(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    folders = db.query(TestCaseFolder).filter(
        TestCaseFolder.user_id == current_user.id
    ).all()
    
    return [{"id": folder.id, "name": folder.name} for folder in folders]


class UpdateTestConfigRequest(BaseModel):
    conversation_id: int
    agent_names: List[str]
    test_case: str


@router.post("/update-test-config")
async def update_test_config(
    request: UpdateTestConfigRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    conversation = db.query(Conversation).filter(
        Conversation.id == request.conversation_id,
        Conversation.user_id == current_user.id
    ).first()
    
    if not conversation:
        raise HTTPException(status_code=404, detail="对话不存在")
    
    valid_agents = validate_agent_names(db, current_user.id, request.agent_names)
    valid_test_case = validate_test_case(db, current_user.id, request.test_case)
    
    valid_agent_names = [a["name"] for a in valid_agents]
    
    test_state = db.query(TestConversationState).filter(
        TestConversationState.conversation_id == request.conversation_id
    ).first()
    
    if test_state:
        test_state.agent_names = ','.join(valid_agent_names)
        test_state.test_case = valid_test_case or ""
    else:
        test_state = TestConversationState(
            conversation_id=request.conversation_id,
            status="start_testing",
            agent_names=','.join(valid_agent_names),
            test_case=valid_test_case or ""
        )
        db.add(test_state)
    
    updated_message_content = f"测试信息已确认。\n\n**智能体**：{', '.join(valid_agent_names)}\n**测试用例**：{valid_test_case}\n"
    
    assistant_messages = db.query(ChatMessage).filter(
        ChatMessage.conversation_id == request.conversation_id,
        ChatMessage.role == "assistant"
    ).order_by(ChatMessage.id).all()
    
    if assistant_messages:
        latest_message = assistant_messages[-1]
        if "测试信息已确认" in latest_message.content or "智能体" in latest_message.content:
            latest_message.content = updated_message_content
            logger.info(f"Updated message content for conversation {request.conversation_id}")
    
    db.commit()
    
    return {
        "success": True,
        "agent_names": valid_agent_names,
        "test_case": valid_test_case or "",
        "updated_message_content": updated_message_content
    }


class StartTestRequest(BaseModel):
    conversation_id: int
    agent_names: List[str]
    test_case_folder_name: str
    request_interval: int = 0


@router.post("/start-test")
async def start_test(
    request: StartTestRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    conversation = db.query(Conversation).filter(
        Conversation.id == request.conversation_id,
        Conversation.user_id == current_user.id
    ).first()
    
    if not conversation:
        raise HTTPException(status_code=404, detail="对话不存在")
    
    agents = db.query(Agent).filter(
        Agent.user_id == current_user.id,
        Agent.name.in_(request.agent_names),
        Agent.is_active == True
    ).all()
    
    workflow_agents = db.query(WorkflowAgent).filter(
        WorkflowAgent.user_id == current_user.id,
        WorkflowAgent.name.in_(request.agent_names),
        WorkflowAgent.is_active == True
    ).all()
    
    if not agents and not workflow_agents:
        raise HTTPException(status_code=400, detail="未找到有效的智能体")
    
    folder = db.query(TestCaseFolder).filter(
        TestCaseFolder.user_id == current_user.id,
        TestCaseFolder.name == request.test_case_folder_name
    ).first()
    
    if not folder:
        raise HTTPException(status_code=400, detail="未找到测试用例文件夹")
    
    all_agent_names = [a.name for a in agents] + [a.name for a in workflow_agents]
    
    test_state = db.query(TestConversationState).filter(
        TestConversationState.conversation_id == request.conversation_id
    ).first()
    
    if test_state:
        test_state.status = "testing"
    else:
        test_state = TestConversationState(
            conversation_id=request.conversation_id,
            status="testing",
            agent_names=','.join(all_agent_names),
            test_case=folder.name,
            request_interval=request.request_interval
        )
        db.add(test_state)
    
    start_message = ChatMessage(
        conversation_id=request.conversation_id,
        role="assistant",
        content="开始测试..."
    )
    db.add(start_message)
    
    db.commit()
    
    agent_ids = [agent.id for agent in agents]
    workflow_agent_ids = [agent.id for agent in workflow_agents]
    
    background_tasks.add_task(
        run_tests_for_multiple_agents,
        agent_ids=agent_ids,
        workflow_agent_ids=workflow_agent_ids,
        folder_id=folder.id,
        user_id=current_user.id,
        conversation_id=request.conversation_id,
        request_interval=request.request_interval
    )
    
    return {
        "success": True,
        "message": "开始测试...",
        "agent_count": len(agents) + len(workflow_agents),
        "test_folder": folder.name
    }


@router.get("/test-results/agents")
async def get_agents_with_test_results(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    agent_results = db.query(
        TestResult.agent_id,
        Agent.name,
        func.count(TestResult.id).label('test_count')
    ).join(
        Agent, TestResult.agent_id == Agent.id
    ).filter(
        Agent.user_id == current_user.id,
        TestResult.agent_type == "agent"
    ).group_by(
        TestResult.agent_id, Agent.name
    ).all()
    
    workflow_agent_results = db.query(
        TestResult.agent_id,
        WorkflowAgent.name,
        func.count(TestResult.id).label('test_count')
    ).join(
        WorkflowAgent, TestResult.agent_id == WorkflowAgent.id
    ).filter(
        WorkflowAgent.user_id == current_user.id,
        TestResult.agent_type == "workflow_agent"
    ).group_by(
        TestResult.agent_id, WorkflowAgent.name
    ).all()
    
    results = []
    for r in agent_results:
        results.append({
            "agent_id": r.agent_id,
            "agent_name": r.name,
            "agent_type": "agent",
            "test_count": r.test_count
        })
    
    for r in workflow_agent_results:
        results.append({
            "agent_id": r.agent_id,
            "agent_name": r.name,
            "agent_type": "workflow_agent",
            "test_count": r.test_count
        })
    
    return results


@router.get("/test-results/{agent_id}/folders")
async def get_test_result_folders(
    agent_id: int,
    agent_type: str = "agent",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if agent_type == "workflow_agent":
        agent = db.query(WorkflowAgent).filter(
            WorkflowAgent.id == agent_id,
            WorkflowAgent.user_id == current_user.id
        ).first()
        
        if not agent:
            raise HTTPException(status_code=404, detail="工作流智能体不存在")
    else:
        agent = db.query(Agent).filter(
            Agent.id == agent_id,
            Agent.user_id == current_user.id
        ).first()
        
        if not agent:
            raise HTTPException(status_code=404, detail="智能体不存在")
    
    folders = db.query(
        TestResult.test_folder_id,
        TestCaseFolder.name,
        func.count(TestResult.id).label('test_count')
    ).join(
        TestCaseFolder, TestResult.test_folder_id == TestCaseFolder.id
    ).filter(
        TestResult.agent_id == agent_id,
        TestResult.agent_type == agent_type
    ).group_by(
        TestResult.test_folder_id, TestCaseFolder.name
    ).all()
    
    return [
        {
            "folder_id": f.test_folder_id,
            "folder_name": f.name,
            "test_count": f.test_count
        }
        for f in folders
    ]


@router.get("/test-results/{agent_id}/folder/{folder_id}")
async def get_test_results_by_folder(
    agent_id: int,
    folder_id: str,
    agent_type: str = "agent",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if agent_type == "workflow_agent":
        agent = db.query(WorkflowAgent).filter(
            WorkflowAgent.id == agent_id,
            WorkflowAgent.user_id == current_user.id
        ).first()
        
        if not agent:
            raise HTTPException(status_code=404, detail="工作流智能体不存在")
    else:
        agent = db.query(Agent).filter(
            Agent.id == agent_id,
            Agent.user_id == current_user.id
        ).first()
        
        if not agent:
            raise HTTPException(status_code=404, detail="智能体不存在")
    
    results = db.query(TestResult).filter(
        TestResult.agent_id == agent_id,
        TestResult.agent_type == agent_type,
        TestResult.test_folder_id == folder_id
    ).order_by(TestResult.created_at).all()
    
    test_details = []
    for result in results:
        test_case = db.query(TestCase).filter(TestCase.id == result.test_case_id).first()
        
        file_info = None
        if result.file_id:
            test_file = db.query(BaseFile).filter(BaseFile.id == result.file_id).first()
            if test_file:
                file_info = {
                    "file_id": test_file.id,
                    "filename": test_file.filename
                }
        
        images = []
        if result.image_ids:
            for image_id in result.image_ids:
                images.append({
                    "image_id": image_id
                })
        
        test_details.append({
            "id": result.id,
            "test_case_id": result.test_case_id,
            "question": result.question or (test_case.question if test_case else ""),
            "sample_answer": test_case.sample_answer if test_case else "",
            "file_info": file_info,
            "images": images,
            "response": result.response,
            "error_message": result.error_message,
            "request_time": result.request_time,
            "created_at": result.created_at.isoformat() if result.created_at else None
        })
    
    return test_details


@router.delete("/test-results/{result_id}")
async def delete_test_result(
    result_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = db.query(TestResult).filter(TestResult.id == result_id).first()
    
    if not result:
        raise HTTPException(status_code=404, detail="测试结果不存在")
    
    if result.agent_type == "workflow_agent":
        agent = db.query(WorkflowAgent).filter(
            WorkflowAgent.id == result.agent_id,
            WorkflowAgent.user_id == current_user.id
        ).first()
    else:
        agent = db.query(Agent).filter(
            Agent.id == result.agent_id,
            Agent.user_id == current_user.id
        ).first()
    
    if not agent:
        raise HTTPException(status_code=404, detail="无权删除此测试结果")
    
    db.delete(result)
    db.commit()
    
    return {"message": "删除成功"}


@router.post("/test-results/batch-delete")
async def batch_delete_test_results(
    result_ids: List[int],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not result_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的测试结果")
    
    deleted_count = 0
    for result_id in result_ids:
        result = db.query(TestResult).filter(TestResult.id == result_id).first()
        
        if not result:
            continue
        
        if result.agent_type == "workflow_agent":
            agent = db.query(WorkflowAgent).filter(
                WorkflowAgent.id == result.agent_id,
                WorkflowAgent.user_id == current_user.id
            ).first()
        else:
            agent = db.query(Agent).filter(
                Agent.id == result.agent_id,
                Agent.user_id == current_user.id
            ).first()
        
        if agent:
            db.delete(result)
            deleted_count += 1
    
    db.commit()
    
    return {"message": f"成功删除 {deleted_count} 条测试结果", "deleted_count": deleted_count}


@router.get("/test-results/{agent_id}/folder/{folder_id}/export")
async def export_test_results_excel(
    agent_id: int,
    folder_id: str,
    agent_type: str = "agent",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if agent_type == "workflow_agent":
        agent = db.query(WorkflowAgent).filter(
            WorkflowAgent.id == agent_id,
            WorkflowAgent.user_id == current_user.id
        ).first()
        
        if not agent:
            raise HTTPException(status_code=404, detail="工作流智能体不存在")
    else:
        agent = db.query(Agent).filter(
            Agent.id == agent_id,
            Agent.user_id == current_user.id
        ).first()
        
        if not agent:
            raise HTTPException(status_code=404, detail="智能体不存在")
    
    folder = db.query(TestCaseFolder).filter(
        TestCaseFolder.id == folder_id
    ).first()
    
    results = db.query(TestResult).filter(
        TestResult.agent_id == agent_id,
        TestResult.agent_type == agent_type,
        TestResult.test_folder_id == folder_id
    ).order_by(TestResult.created_at).all()
    
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=500, detail="无法创建工作表")
    ws.title = "测试结果"
    
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    headers = ["序号", "问题", "图片", "文件", "智能体回答", "预设回答", "错误信息", "耗时(s)"]
    col_widths = [6, 40, 25, 20, 50, 50, 30, 10]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if cell:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    for row_idx, result in enumerate(results, 2):
        test_case = db.query(TestCase).filter(TestCase.id == result.test_case_id).first()
        
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='top')
        
        question_cell = ws.cell(row=row_idx, column=2, value=result.question or (test_case.question if test_case else ""))
        question_cell.border = thin_border
        question_cell.alignment = wrap_alignment
        
        image_cell = ws.cell(row=row_idx, column=3, value="")
        image_cell.border = thin_border
        image_cell.alignment = wrap_alignment
        
        max_img_height = 60
        
        if result.image_ids:
            img_offset = 0
            for img_idx, image_id in enumerate(result.image_ids, 1):
                try:
                    extracted_img = db.query(ExtractedImage).filter(
                        ExtractedImage.id == image_id
                    ).first()
                    
                    if extracted_img and extracted_img.image_path:
                        image_data = file_storage.read_image(extracted_img.image_path)
                        img_bytes = BytesIO(image_data)
                        pil_img = PILImage.open(img_bytes)
                        
                        max_width = 150
                        max_height = 100
                        pil_img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                        
                        img_buffer = BytesIO()
                        pil_img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        xl_img = XLImage(img_buffer)
                        xl_img.anchor = f'C{row_idx}'
                        ws.add_image(xl_img)
                        
                        img_height = min(pil_img.height + 10, 120)
                        if img_height > max_img_height:
                            max_img_height = img_height
                        img_offset += pil_img.height + 5
                except Exception as e:
                    logger.error(f"Failed to insert image {image_id}: {e}")
        
        file_cell = ws.cell(row=row_idx, column=4, value="")
        file_cell.border = thin_border
        file_cell.alignment = wrap_alignment
        
        if result.file_id:
            test_file = db.query(BaseFile).filter(BaseFile.id == result.file_id).first()
            if test_file:
                file_cell.value = f"{test_file.filename}"
        
        response_cell = ws.cell(row=row_idx, column=5, value=result.response or "无响应")
        response_cell.border = thin_border
        response_cell.alignment = wrap_alignment
        
        sample_cell = ws.cell(row=row_idx, column=6, value=test_case.sample_answer if test_case else "")
        sample_cell.border = thin_border
        sample_cell.alignment = wrap_alignment
        
        error_cell = ws.cell(row=row_idx, column=7, value=result.error_message or "")
        error_cell.border = thin_border
        error_cell.alignment = wrap_alignment
        
        time_cell = ws.cell(row=row_idx, column=8, value=round(result.request_time, 2) if result.request_time else 0)
        time_cell.border = thin_border
        time_cell.alignment = Alignment(horizontal='center', vertical='top')
        
        ws.row_dimensions[row_idx].height = max_img_height
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    folder_name = folder.name if folder else folder_id
    agent_name = agent.name
    filename = f"测试结果_{agent_name}_{folder_name}.xlsx"
    encoded_filename = quote(filename)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.websocket("/ws/{user_id}")
async def websocket_endpoint(websocket: WebSocket, user_id: int):
    await manager.connect(websocket, user_id)
    try:
        while True:
            data = await websocket.receive_text()
            try:
                message = json.loads(data)
                if message.get("type") == "ping":
                    await websocket.send_text(json.dumps({"type": "pong"}))
            except json.JSONDecodeError:
                pass
    except WebSocketDisconnect:
        await manager.disconnect(websocket, user_id)


@router.post("/generate-test-case", response_model=GenerateTestCaseResponse)
async def generate_test_case(
    request: GenerateTestCaseRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    user_id = current_user.id
    agent_names = request.agent_names
    user_requirement = request.requirement or ""
    count = request.count or 5
    folder_id = request.folder_id
    conversation_id = None

    background_tasks.add_task(
        generate_test_case_background,
        agent_names=agent_names,
        user_id=user_id,
        user_requirement=user_requirement,
        count=count,
        folder_id=folder_id,
        conversation_id=conversation_id
    )
    
    # 返回一个空的测试用例列表，因为测试用例是在后台生成的
    return GenerateTestCaseResponse(
        test_cases=[],
        agent_names=agent_names
    )


async def generate_test_case_background(
    agent_names: List[str],
    user_id: int,
    user_requirement: str = "",
    count: int = 5,
    folder_id: Optional[str] = None,
    conversation_id: Optional[int] = None
):
    db = SessionLocal()
    
    try:
        logger.info(f"=== 后台任务：开始生成测试用例 ===")
        logger.info(f"智能体名称: {agent_names}")
        
        await manager.send_to_user(user_id, {
            "type": "test_case_generation_started",
            "agent_names": agent_names,
            "message": f"开始为智能体 {'、'.join(agent_names)} 生成测试用例..."
        })
        
        agents = db.query(Agent).filter(
            Agent.user_id == user_id,
            Agent.name.in_(agent_names),
            Agent.is_active == True
        ).all()
        
        workflow_agents = db.query(WorkflowAgent).filter(
            WorkflowAgent.user_id == user_id,
            WorkflowAgent.name.in_(agent_names),
            WorkflowAgent.is_active == True
        ).all()
        
        if not agents and not workflow_agents:
            await manager.send_to_user(user_id, {
                "type": "test_case_generation_error",
                "agent_names": agent_names,
                "message": "未找到有效的智能体"
            })
            return
        
        agents_info = []
        for agent in agents:
            agent_info = f"智能体名称: {agent.name}"
            if agent.description:
                agent_info += f"\n智能体描述: {agent.description}"
            agents_info.append(agent_info)
        
        agents_info_str = "\n\n".join(agents_info)
        
        from app.services.functional_config import get_code_gen_config
        
        code_gen_model = get_code_gen_config(db)
        
        if code_gen_model:
            model_code = code_gen_model.get("model", "")
            model_api_key = code_gen_model.get("api_key", "")
            model_api_url = code_gen_model.get("url", "https://api.openai.com/v1")
            call_type = str(code_gen_model.get("call_type", "OpenAI Chat"))
        else:
            default_config = db.query(ApiConfig).filter(
                ApiConfig.user_id == user_id,
                ApiConfig.is_default == True
            ).first()
            
            if not default_config:
                default_config = db.query(ApiConfig).filter(
                    ApiConfig.user_id == user_id
                ).first()
            
            if not default_config:
                await manager.send_to_user(user_id, {
                    "type": "test_case_generation_error",
                    "agent_names": agent_names,
                    "message": "请先配置默认大模型或由管理员配置代码生成模型"
                })
                return
            
            model_code = default_config.code
            model_api_key = default_config.api_key
            model_api_url = default_config.api_url
            call_type = str(default_config.call_type) if default_config.call_type is not None else "OpenAI Chat"
        
        await manager.send_to_user(user_id, {
            "type": "test_case_generation_progress",
            "agent_names": agent_names,
            "message": "正在调用AI模型生成测试用例..."
        })

        try:
            url = str(model_api_url) if model_api_url else "https://api.anthropic.com"

            messages = [
                {"role": "system", "content": TEST_CASE_GENERATION_SYSTEM_PROMPT},
                {"role": "user", "content": TEST_CASE_GENERATION_USER_PROMPT.format(
                    agents_info=agents_info_str,
                    user_requirement=user_requirement if user_requirement else "无特殊要求，请根据智能体功能生成合适的测试用例",
                    count=count
                )}
            ]

            result = await call_llm(
                messages=messages,
                model=model_code,
                api_key=model_api_key,
                url=url,
                call_type=call_type,
                structured_output=False,
                timeout=60.0,
                default_response={"content": "[]"}
            )

            generated_content = result.get("content", "")
            generated_content = str(generated_content) if generated_content else ""
            json_match = re.search(r'```json\s*(.*?)\s*```', generated_content, re.DOTALL)
            if json_match:
                generated_content = json_match.group(1)
            else:
                json_match = re.search(r'\[\s*\{.*?\}\s*\]', generated_content, re.DOTALL)
                if json_match:
                    generated_content = json_match.group(0)

            test_cases_data = json.loads(generated_content)
                
        except json.JSONDecodeError as e:
            logger.error(f"JSON解析失败: {str(e)}")
            await manager.send_to_user(user_id, {
                "type": "test_case_generation_error",
                "agent_names": agent_names,
                "message": "生成的测试用例格式不正确"
            })
            return
        except Exception as e:
            logger.error(f"生成测试用例失败: {str(e)}")
            await manager.send_to_user(user_id, {
                "type": "test_case_generation_error",
                "agent_names": agent_names,
                "message": f"生成测试用例失败: {str(e)}"
            })
            return
        
        await manager.send_to_user(user_id, {
            "type": "test_case_generation_progress",
            "agent_names": agent_names,
            "message": f"成功生成 {len(test_cases_data)} 个测试用例，正在保存..."
        })
        
        # 使用指定的文件夹或创建新文件夹
        if folder_id:
            folder = db.query(TestCaseFolder).filter(
                TestCaseFolder.id == folder_id,
                TestCaseFolder.user_id == user_id
            ).first()
            if not folder:
                await manager.send_to_user(user_id, {
                    "type": "test_case_generation_error",
                    "agent_names": agent_names,
                    "message": "指定的测试用例文件夹不存在或无权限"
                })
                return
        else:
            folder_name = f"{'、'.join(agent_names)}的测试用例-{int(time.time())}"
            folder = db.query(TestCaseFolder).filter(
                TestCaseFolder.user_id == user_id,
                TestCaseFolder.name == folder_name
            ).first()
            
            if not folder:
                folder = TestCaseFolder(
                    name=folder_name,
                    user_id=user_id,
                    description=f"由智能体 {'、'.join(agent_names)} 自动生成的测试用例",
                    case_count=0
                )
                db.add(folder)
                db.flush()
                logger.info(f"创建测试用例文件夹: {folder_name}")
        
        max_row_order_result = db.query(func.max(TestCase.row_order)).filter(
            TestCase.folder_id == folder.id
        ).scalar()
        max_row_order = max_row_order_result if max_row_order_result is not None else -1
        
        for idx, tc in enumerate(test_cases_data):
            test_case = TestCase(
                folder_id=folder.id,
                user_id=user_id,
                question=tc.get("question", ""),
                sample_answer=tc.get("sample_answer", ""),
                row_order=max_row_order + idx + 1
            )
            db.add(test_case)
        
        folder.case_count = len(test_cases_data)
        db.commit()
        
        logger.info(f"=== 测试用例生成完成 ===")
        logger.info(f"保存了 {len(test_cases_data)} 个测试用例到文件夹: {folder.name}")
        
        await manager.send_to_user(user_id, {
            "type": "test_case_generation_completed",
            "agent_names": agent_names,
            "folder_name": folder.name,
            "test_case_count": len(test_cases_data),
            "message": f"成功生成并保存 {len(test_cases_data)} 个测试用例到文件夹「{folder.name}」"
        })
        
        if conversation_id:
            test_state = db.query(TestConversationState).filter(
                TestConversationState.conversation_id == conversation_id
            ).first()
            if test_state:
                test_state.status = "success_generating"
                test_state.test_case = folder_name
                db.commit()
        
    except Exception as e:
        logger.error(f"后台任务执行失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        
        await manager.send_to_user(user_id, {
            "type": "test_case_generation_error",
            "agent_names": agent_names,
            "message": f"生成测试用例失败: {str(e)}"
        })
        
        if conversation_id:
            test_state = db.query(TestConversationState).filter(
                TestConversationState.conversation_id == conversation_id
            ).first()
            if test_state:
                test_state.status = "failure_generating"
                db.commit()
    finally:
        db.close()
